import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def build_report_excel(data: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт"
 
    # --- Стилі ---
    header_font   = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell_font     = Font(name="Arial", size=11)
    header_fill   = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    alt_fill      = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    center_align  = Alignment(horizontal="center", vertical="center")
    thin_side     = Side(style="thin", color="000000")
    thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
 
    # --- Заголовок таблиці ---
    # 1: Клас, 2: Всього, 3: Відсутні, 4: Хворі
    headers = ["Клас", "Всього учнів у класі", "Відсутні", "Хворих"]
    col_widths = [10, 25, 14, 12]
 
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = width
 
    ws.row_dimensions[1].height = 20
 
    # --- Рядки даних ---
    for row_idx, item in enumerate(data, start=2):
        # Оновлений порядок даних згідно з твоїм запитом
        row_data = [
            item["class"],   # 1 колонка
            item["total"],   # 2 колонка
            item["absent"],  # 3 колонка
            item["sick"]     # 4 колонка
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
 
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = cell_font
            cell.alignment = center_align
            cell.border    = thin_border
            if fill:
                cell.fill = fill
 
    # --- Підсумковий рядок ---
    summary_row = len(data) + 2
    
    # Комірка "Разом"
    summary_label = ws.cell(row=summary_row, column=1, value="Разом")
    summary_label.font = header_font
    summary_label.fill = header_fill
    summary_label.alignment = center_align
    summary_label.border = thin_border
 
    # Формули SUM для колонок 2, 3 та 4
    for col_idx in range(2, 5):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        cell = ws.cell(
            row=summary_row,
            column=col_idx,
            value=f"=SUM({col_letter}2:{col_letter}{summary_row - 1})",
        )
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
 
    # --- Зберегти у байти ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
